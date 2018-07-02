# processInterns.py
import sys
from openpyxl import *
from intern import Intern
from collections import OrderedDict

dataset = []

def loadExcel(raw_data):
    global dataset
    workbook = load_workbook(raw_data, data_only=True)
    spreadsheet = workbook["InternData"]
    dataset = spreadsheet
    return spreadsheet

def createInterns():
    interns = []
    count = 1
    for intern, in dataset.iter_rows(max_col=1, max_row = 9):
        newIntern = Intern(count, intern.value)
        interns.append(newIntern)
        count += 1
    return interns

def getAttributes(internRow):
    att_list = []
    for cell, in dataset.iter_cols(min_row=internRow, max_row=internRow, min_col=2, max_col=5):
        att_list.append(cell.value)
    attributes_tuple = (att_list[0], att_list[1], att_list[2], att_list[3])
    return attributes_tuple

def setPairs(interns):
    for intern in interns:
        setEndorsers(intern, interns)
        rankEndorsers(intern)
        pruneEndorsers(intern)
        # pairEndorsers(interns)

def setEndorsers(intern, interns):
    for other_intern in interns:
        if intern.name != other_intern.name:
            score = scorePair(intern, other_intern)
            intern.endorsers.update({other_intern.name: score})

# compare intern attributes
def scorePair(intern, other_intern):
    compatibility_score = 0
    attributes = ["team", "position", "product", "skills"]
    matching_words = []
    for attribute in attributes:
        for word in getattr(intern, attribute).split():
            if word in getattr(other_intern, attribute).split():
                compatibility_score += 1
    return compatibility_score

# order endorsers by quantity of matches
def rankEndorsers(intern):
    intern.endorsers = OrderedDict(sorted(intern.endorsers.items(), key=lambda kv: kv[1], reverse=True))

# remove endorsers with no matches
def pruneEndorsers(intern):
    intern.endorsers = {key: value for key, value in intern.endorsers.items() if value is not 0}

# implement stable marriage algo
# https://en.wikipedia.org/wiki/Stable_marriage_problem#Algorithm
def pairEndorsers(interns):
    rounds = len(interns) / 2
    for round in range(rounds):
        marry(interns)

# def marry(interns):
#     pairs = {}
#     for intern in interns:
#         pairs.update({intern: 0})
#     while 0 in pairs.values():
#       print (deez nyets)

def getPairs(interns):
    for intern in interns:
        print (intern.name.upper())
        for key, value in intern.endorsers.items():
            print (key + "\t" + str(value))
        print ("\n")

def checkData(table):
    names = []
    for intern, in dataset.iter_rows(max_col=1, max_row = 9):
        names += intern.value
    for name in names:
        for name_compare in names:
            matching_names = 0
            if name == name_compare:
                matching_names += 1
            if matching_names > 1:
                sys.exit(matching_names, "instances of", name)
